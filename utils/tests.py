from django.test import TestCase, RequestFactory
from contacts.models import Contact
from utils.search import filter_queryset, get_search_query, ModelSearcher


class ReusableSearchTestCase(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.c1 = Contact.objects.create(
            first_name='Alice', last_name='Smith', email='alice@example.com', tag='VIP'
        )
        self.c2 = Contact.objects.create(
            first_name='Bob', last_name='Jones', email='bob@test.org', tag='Lead'
        )
        self.c3 = Contact.objects.create(
            first_name='Charlie', last_name='Smith', email='charlie@enterprise.com', tag='Customer'
        )

    def test_get_search_query_normalizes_input(self):
        request = self.factory.get('/?q=  john doe  ')
        self.assertEqual(get_search_query(request, 'q'), 'john doe')

        request_empty = self.factory.get('/')
        self.assertEqual(get_search_query(request_empty, 'q'), '')

    def test_filter_queryset_single_term(self):
        qs = Contact.objects.all()
        fields = ['first_name', 'last_name', 'email', 'tag']

        results = filter_queryset(qs, 'Alice', fields)
        self.assertEqual(list(results), [self.c1])

        results_tag = filter_queryset(qs, 'Lead', fields)
        self.assertEqual(list(results_tag), [self.c2])

    def test_filter_queryset_case_insensitive(self):
        qs = Contact.objects.all()
        fields = ['first_name', 'last_name', 'email', 'tag']

        results = filter_queryset(qs, 'smith', fields)
        self.assertCountEqual(list(results), [self.c1, self.c3])

    def test_filter_queryset_multi_word_search(self):
        qs = Contact.objects.all()
        fields = ['first_name', 'last_name', 'email', 'tag']

        # 'Charlie Smith' should match Charlie Smith
        results = filter_queryset(qs, 'Charlie Smith', fields)
        self.assertEqual(list(results), [self.c3])

        # 'Smith VIP' should match Alice Smith who has tag VIP
        results_vip = filter_queryset(qs, 'Smith VIP', fields)
        self.assertEqual(list(results_vip), [self.c1])

    def test_filter_queryset_empty_or_none(self):
        qs = Contact.objects.all()
        fields = ['first_name', 'last_name']

        self.assertEqual(filter_queryset(qs, '', fields).count(), 3)
        self.assertEqual(filter_queryset(qs, '   ', fields).count(), 3)

    def test_model_searcher_class(self):
        searcher = ModelSearcher(['first_name', 'last_name', 'email', 'tag'])
        request = self.factory.get('/contacts/?q=bob')

        filtered_qs, query_str = searcher.search(request, Contact.objects.all())
        self.assertEqual(query_str, 'bob')
        self.assertEqual(list(filtered_qs), [self.c2])


import io
import openpyxl
from utils.importer import ModelFileImporter, parse_csv_file, parse_excel_file, normalize_header


class ReusableImporterTestCase(TestCase):
    def setUp(self):
        self.importer = ModelFileImporter(
            model_class=Contact,
            field_mapping={
                'first_name': ['first_name', 'first name', 'fname'],
                'last_name': ['last_name', 'last name', 'lname'],
                'email': ['email', 'email address', 'e-mail'],
                'tag': ['tag', 'tags', 'group'],
            },
            required_fields=['first_name', 'email'],
            unique_key='email',
            update_existing=True,
        )

    def test_normalize_header(self):
        self.assertEqual(normalize_header(' First Name '), 'firstname')
        self.assertEqual(normalize_header('E-Mail_Address'), 'emailaddress')

    def test_parse_csv_file(self):
        csv_data = "First Name,Last Name,Email,Tag\nJohn,Doe,john@example.com,VIP\nJane,Smith,jane@test.org,Customer"
        file_obj = io.BytesIO(csv_data.encode('utf-8'))
        rows = parse_csv_file(file_obj)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]['First Name'], 'John')
        self.assertEqual(rows[1]['Email'], 'jane@test.org')

    def test_parse_excel_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['First Name', 'Last Name', 'Email', 'Tag'])
        ws.append(['Alex', 'Ray', 'alex@excel.com', 'Lead'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        rows = parse_excel_file(output)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['First Name'], 'Alex')
        self.assertEqual(rows[0]['Email'], 'alex@excel.com')

    def test_import_rows_create_and_update(self):
        rows = [
            {'first_name': 'Mark', 'last_name': 'Taylor', 'email': 'mark@test.com', 'tag': 'Lead'},
            {'first_name': 'Sarah', 'last_name': 'Connor', 'email': 'sarah@test.com', 'tag': 'VIP'},
        ]
        result = self.importer.import_rows(rows)
        self.assertEqual(result.created, 2)
        self.assertEqual(Contact.objects.count(), 2)

        rows_update = [
            {'first_name': 'Mark', 'last_name': 'Taylor Updated', 'email': 'mark@test.com', 'tag': 'Customer'},
        ]
        result_update = self.importer.import_rows(rows_update)
        self.assertEqual(result_update.updated, 1)
        self.assertEqual(Contact.objects.get(email='mark@test.com').last_name, 'Taylor Updated')


from utils.exporter import ModelFileExporter, export_to_csv_response, export_to_excel_response


class ReusableExporterTestCase(TestCase):
    def setUp(self):
        self.c1 = Contact.objects.create(
            first_name='Alice', last_name='Smith', email='alice@example.com', tag='VIP'
        )
        self.exporter = ModelFileExporter(
            fields=['first_name', 'last_name', 'email', 'tag'],
            header_labels={'first_name': 'First Name', 'last_name': 'Last Name'},
            default_filename='test_export',
        )

    def test_export_csv_response(self):
        response = self.exporter.export_response(Contact.objects.all(), format='csv')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')
        self.assertIn('attachment; filename="test_export.csv"', response['Content-Disposition'])
        content = response.content.decode('utf-8')
        self.assertIn('First Name,Last Name,Email,Tag', content)
        self.assertIn('Alice,Smith,alice@example.com,VIP', content)

    def test_export_excel_response(self):
        response = self.exporter.export_response(Contact.objects.all(), format='excel')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment; filename="test_export.xlsx"', response['Content-Disposition'])
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0], ('First Name', 'Last Name', 'Email', 'Tag'))
        self.assertEqual(rows[1], ('Alice', 'Smith', 'alice@example.com', 'VIP'))


