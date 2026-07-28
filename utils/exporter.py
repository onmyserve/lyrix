import csv
import io
from typing import Sequence, Dict, Optional, Any, Iterable, Union
from django.http import HttpResponse


def export_to_csv_response(
    data: Iterable[Union[Dict[str, Any], Any]],
    fields: Sequence[str],
    filename: str = 'export.csv',
    header_labels: Optional[Dict[str, str]] = None,
) -> HttpResponse:
    """
    Exports a QuerySet or list of model instances/dicts into a downloadable CSV HttpResponse.
    """
    header_labels = header_labels or {}
    headers = [header_labels.get(f, f.replace('_', ' ').title()) for f in fields]

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    if not filename.endswith('.csv'):
        filename += '.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(headers)

    for item in data:
        row = []
        for field_name in fields:
            if isinstance(item, dict):
                val = item.get(field_name, '')
            else:
                val = getattr(item, field_name, '')
                if callable(val):
                    val = val()
            row.append(str(val) if val is not None else '')
        writer.writerow(row)

    return response


def export_to_excel_response(
    data: Iterable[Union[Dict[str, Any], Any]],
    fields: Sequence[str],
    filename: str = 'export.xlsx',
    header_labels: Optional[Dict[str, str]] = None,
    sheet_name: str = 'Export',
) -> HttpResponse:
    """
    Exports a QuerySet or list of model instances/dicts into a formatted Excel (.xlsx) HttpResponse.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl is required to export Excel files.") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_labels = header_labels or {}
    headers = [header_labels.get(f, f.replace('_', ' ').title()) for f in fields]

    # Style header row
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_align = Alignment(horizontal='left', vertical='center')

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for item in data:
        row = []
        for field_name in fields:
            if isinstance(item, dict):
                val = item.get(field_name, '')
            else:
                val = getattr(item, field_name, '')
                if callable(val):
                    val = val()
            row.append(str(val) if val is not None else '')
        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class ModelFileExporter:
    """
    Reusable Model Exporter class for Django models.

    Usage:
        exporter = ModelFileExporter(
            fields=['first_name', 'last_name', 'email', 'tag', 'created_at'],
            header_labels={'first_name': 'First Name', 'last_name': 'Last Name', 'email': 'Email'},
            default_filename='contacts'
        )
        return exporter.export_response(queryset, format='csv')
    """

    def __init__(
        self,
        fields: Sequence[str],
        header_labels: Optional[Dict[str, str]] = None,
        default_filename: str = 'export',
        sheet_name: str = 'Sheet1',
    ):
        self.fields = list(fields)
        self.header_labels = header_labels or {}
        self.default_filename = default_filename
        self.sheet_name = sheet_name

    def export_response(
        self,
        data: Iterable[Union[Dict[str, Any], Any]],
        format: str = 'csv',
        filename: Optional[str] = None,
    ) -> HttpResponse:
        fn = filename or self.default_filename
        fmt = str(format).lower().strip()

        if fmt in ('xlsx', 'excel'):
            return export_to_excel_response(
                data=data,
                fields=self.fields,
                filename=f"{fn}.xlsx" if not fn.endswith('.xlsx') else fn,
                header_labels=self.header_labels,
                sheet_name=self.sheet_name,
            )
        else:
            return export_to_csv_response(
                data=data,
                fields=self.fields,
                filename=f"{fn}.csv" if not fn.endswith('.csv') else fn,
                header_labels=self.header_labels,
            )
